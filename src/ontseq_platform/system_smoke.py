from __future__ import annotations

import os
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook
from pydantic import Field

from .cnv.extension import QDNAseqExtensionSettings, register_qdnaseq_extension
from .cnv.qdnaseq import QDNAseqCallReport, QDNAseqPolicy
from .execution import StreamingCommandRunner, SubprocessRunner
from .io import write_json
from .models import (
    AnalysisIntent,
    AnalysisModule,
    AnalysisSpec,
    AssayMode,
    AssaySpec,
    CheckStatus,
    EventType,
    GenomeBuild,
    InputKind,
    InputSpec,
    ModuleRunStatus,
    PipelineResult,
    QCPolicy,
    ReferenceLock,
    SampleManifest,
    SnifflesPolicy,
    StrictModel,
    ValidationCheck,
    Verdict,
)
from .pipeline.envelope import sha256_file
from .pipeline.runner import RunConfiguration, run_pipeline
from .pipeline.stages import StageId
from .reference import reference_lock_from_fai
from .smoke import run_local_smoke

SYSTEM_SMOKE_SAMPLE_ID = "SYSTEM_CNV_SMOKE_001"
SYSTEM_SMOKE_RUN_ID = "SYSTEM_CNV_SMOKE_RUN_001"
SYSTEM_SMOKE_REFERENCE_ID = "SYNTHETIC_HG19_CNV_SYSTEM_SMOKE"

# UCSC hg19 / GRCh37 autosome lengths. These are intentionally identical to the
# real-tool CNV CI fixture that already exercises QDNAseq.hg19 + ACE.
GRCH37_AUTOSOME_LENGTHS: dict[int, int] = {
    1: 249250621,
    2: 243199373,
    3: 198022430,
    4: 191154276,
    5: 180915260,
    6: 171115067,
    7: 159138663,
    8: 146364022,
    9: 141213431,
    10: 135534747,
    11: 135006516,
    12: 133851895,
    13: 115169878,
    14: 107349540,
    15: 102531392,
    16: 90354753,
    17: 81195210,
    18: 78077248,
    19: 59128983,
    20: 63025520,
    21: 48129895,
    22: 51304566,
}


class SystemSmokeReport(StrictModel):
    """Compact evidence that the installed runtime executed its critical local lanes."""

    schema_version: Literal["0.1.0"] = "0.1.0"
    verdict: Verdict
    checks: list[ValidationCheck] = Field(min_length=1)
    output_paths: dict[str, str]
    limitations: list[str] = Field(default_factory=list)


def synthetic_cnv_copy_count(chromosome: int) -> int:
    """Return deterministic synthetic read-depth multiplicity for one chromosome."""

    if chromosome not in GRCH37_AUTOSOME_LENGTHS:
        raise ValueError(f"unsupported synthetic chromosome: {chromosome}")
    if chromosome == 7:
        return 1
    if chromosome == 8:
        return 3
    return 2


def _run_checked(
    runner: StreamingCommandRunner,
    argv: list[str],
    *,
    label: str,
    timeout_seconds: int,
) -> None:
    result = runner.run(argv, timeout_seconds=timeout_seconds)
    if result.returncode != 0:
        diagnostic = result.stderr.strip()[-3000:]
        raise ValueError(
            f"{label} failed with exit code {result.returncode}"
            + (f": {diagnostic}" if diagnostic else "")
        )


def _write_cnv_sam(path: Path) -> None:
    sequence = "ACGT" * 25
    quality = "I" * 100
    serial = 0
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write("@HD\tVN:1.6\tSO:unsorted\n")
        for chromosome, length in GRCH37_AUTOSOME_LENGTHS.items():
            handle.write(f"@SQ\tSN:chr{chromosome}\tLN:{length}\n")
        handle.write(f"@RG\tID:SYSTEM_CNV_SMOKE\tSM:{SYSTEM_SMOKE_SAMPLE_ID}\tPL:ONT\n")
        for chromosome, length in GRCH37_AUTOSOME_LENGTHS.items():
            copies = synthetic_cnv_copy_count(chromosome)
            for position in range(50_001, length - 100, 100_000):
                for replicate in range(copies):
                    serial += 1
                    handle.write(
                        f"CNVREAD_{serial:07d}\t0\tchr{chromosome}\t"
                        f"{position + replicate * 5}\t60\t100M\t*\t0\t0\t"
                        f"{sequence}\t{quality}\tRG:Z:SYSTEM_CNV_SMOKE\t"
                        "NM:i:0\tAS:i:100\tde:f:0.0\n"
                    )


def _build_cnv_fixture(
    root: Path,
    *,
    runner: StreamingCommandRunner,
    samtools: str,
    threads: int,
) -> tuple[SampleManifest, ReferenceLock]:
    root.mkdir(parents=True, exist_ok=True)
    sam = root / "synthetic.hg19.sam"
    unsorted = root / "synthetic.hg19.unsorted.bam"
    staged_bam = root / ".synthetic.hg19.bam.tmp"
    bam = root / "synthetic.hg19.bam"
    bai = root / "synthetic.hg19.bam.bai"
    fai = root / "synthetic.hg19.fai"
    manifest_path = root / "sample.manifest.json"
    lock_path = root / "reference.lock.json"

    fixture_targets = (sam, unsorted, staged_bam, bam, bai, fai, manifest_path, lock_path)
    existing = [path.name for path in fixture_targets if path.exists()]
    if existing:
        raise ValueError(
            "Refusing to overwrite existing system-smoke CNV fixture artifacts: "
            + ", ".join(existing)
        )

    _write_cnv_sam(sam)
    fai.write_text(
        "".join(
            f"chr{chromosome}\t{length}\t0\t0\t0\n"
            for chromosome, length in GRCH37_AUTOSOME_LENGTHS.items()
        ),
        encoding="utf-8",
    )

    converted = runner.run_to_file(
        [samtools, "view", "-b", str(sam)],
        unsorted,
        timeout_seconds=300,
    )
    if converted.returncode != 0:
        diagnostic = converted.stderr.strip()[-3000:]
        raise ValueError(
            "samtools BAM conversion failed with exit code "
            f"{converted.returncode}" + (f": {diagnostic}" if diagnostic else "")
        )
    _run_checked(
        runner,
        [samtools, "sort", "-@", str(threads), "-o", str(staged_bam), str(unsorted)],
        label="samtools coordinate sort for CNV smoke",
        timeout_seconds=600,
    )
    os.replace(staged_bam, bam)
    _run_checked(
        runner,
        [samtools, "index", "-@", str(threads), str(bam), str(bai)],
        label="samtools index for CNV smoke",
        timeout_seconds=300,
    )
    _run_checked(
        runner,
        [samtools, "quickcheck", "-v", str(bam)],
        label="samtools quickcheck for CNV smoke",
        timeout_seconds=120,
    )
    sam.unlink(missing_ok=True)
    unsorted.unlink(missing_ok=True)

    reference_lock = reference_lock_from_fai(
        fai,
        reference_id=SYSTEM_SMOKE_REFERENCE_ID,
        genome_build=GenomeBuild.GRCH37,
    )
    manifest = SampleManifest(
        sample_id=SYSTEM_SMOKE_SAMPLE_ID,
        run_id=SYSTEM_SMOKE_RUN_ID,
        input=InputSpec(
            kind=InputKind.ALIGNED_BAM,
            path=str(bam.resolve()),
            index_path=str(bai.resolve()),
            sha256=sha256_file(bam),
        ),
        assay=AssaySpec(
            mode=AssayMode.LOW_COVERAGE_WGS,
            genome_build=GenomeBuild.GRCH37,
            reference_id=SYSTEM_SMOKE_REFERENCE_ID,
        ),
        analysis=AnalysisSpec(
            profile="installed-runtime-system-smoke",
            modules=[AnalysisModule.QC, AnalysisModule.CNV, AnalysisModule.REPORT],
            intent=AnalysisIntent.SOMATIC,
        ),
    )
    write_json(manifest, manifest_path)
    write_json(reference_lock, lock_path)
    return manifest, reference_lock


def cnv_truth_checks(report: QDNAseqCallReport, policy: QDNAseqPolicy) -> list[ValidationCheck]:
    """Validate deterministic chr7-loss / chr8-gain truth against normalized output."""

    observed_bins = sorted(item.bin_size_kbp for item in report.fits)
    expected_bins = sorted(policy.bin_sizes_kbp)
    checks: list[ValidationCheck] = [
        ValidationCheck(
            name="qdnaseq_multibin_execution",
            status=CheckStatus.PASS if observed_bins == expected_bins else CheckStatus.FAIL,
            message=(
                "QDNAseq produced every configured bin-size fit."
                if observed_bins == expected_bins
                else "QDNAseq did not produce exactly the configured bin-size fits."
            ),
            details={
                "expected_bins_kbp": ",".join(str(item) for item in expected_bins),
                "observed_bins_kbp": ",".join(str(item) for item in observed_bins),
            },
        ),
        ValidationCheck(
            name="qdnaseq_primary_bin",
            status=(
                CheckStatus.PASS
                if report.primary_fit.bin_size_kbp == policy.primary_bin_size_kbp
                else CheckStatus.FAIL
            ),
            message=(
                "The configured primary QDNAseq resolution was retained."
                if report.primary_fit.bin_size_kbp == policy.primary_bin_size_kbp
                else "The QDNAseq primary resolution does not match the configured policy."
            ),
            details={
                "expected_primary_bin_kbp": policy.primary_bin_size_kbp,
                "observed_primary_bin_kbp": report.primary_fit.bin_size_kbp,
            },
        ),
    ]

    cellularity_ok = all(0.05 <= fit.cellularity <= 1.0 for fit in report.fits)
    checks.append(
        ValidationCheck(
            name="ace_fit_domain",
            status=CheckStatus.PASS if cellularity_ok else CheckStatus.FAIL,
            message=(
                "ACE returned cellularity estimates in the expected fractional domain."
                if cellularity_ok
                else "ACE returned a cellularity outside the expected 0.05-1.0 domain."
            ),
            details={
                "primary_cellularity": report.primary_fit.cellularity,
                "primary_ploidy": report.primary_fit.ploidy,
                "primary_fit_error": report.primary_fit.fit_error,
            },
        )
    )

    loss7 = [
        event
        for event in report.events
        if event.primary.chromosome == "chr7"
        and event.event_type == EventType.CHROMOSOME_LOSS
        and event.copy_number is not None
        and 0.8 <= event.copy_number <= 1.2
    ]
    gain8 = [
        event
        for event in report.events
        if event.primary.chromosome == "chr8"
        and event.event_type == EventType.CHROMOSOME_GAIN
        and event.copy_number is not None
        and 2.8 <= event.copy_number <= 3.2
    ]
    checks.append(
        ValidationCheck(
            name="expected_chr7_loss",
            status=CheckStatus.PASS if loss7 else CheckStatus.FAIL,
            message=(
                "The synthetic half-depth chromosome 7 was normalized as CN~1 loss."
                if loss7
                else "The expected synthetic chromosome 7 loss was not recovered."
            ),
            details={"matching_events": len(loss7)},
        )
    )
    checks.append(
        ValidationCheck(
            name="expected_chr8_gain",
            status=CheckStatus.PASS if gain8 else CheckStatus.FAIL,
            message=(
                "The synthetic 1.5x-depth chromosome 8 was normalized as CN~3 gain."
                if gain8
                else "The expected synthetic chromosome 8 gain was not recovered."
            ),
            details={"matching_events": len(gain8)},
        )
    )

    consensus = {item.chromosome: item for item in report.chromosome_consensus}
    chr7 = consensus.get("chr7")
    chr8 = consensus.get("chr8")
    expected_agreement = len(policy.bin_sizes_kbp)
    consensus_ok = (
        chr7 is not None
        and chr8 is not None
        and chr7.rounded_copy_number == 1
        and chr8.rounded_copy_number == 3
        and chr7.agreeing_bins == expected_agreement
        and chr8.agreeing_bins == expected_agreement
    )
    checks.append(
        ValidationCheck(
            name="cnv_multibin_consensus",
            status=CheckStatus.PASS if consensus_ok else CheckStatus.FAIL,
            message=(
                "All configured QDNAseq resolutions agree on chr7 CN=1 and chr8 CN=3."
                if consensus_ok
                else "The expected chromosome-level multi-bin CNV consensus was not recovered."
            ),
            details={
                "expected_agreeing_bins": expected_agreement,
                "chr7_cn": chr7.rounded_copy_number if chr7 else None,
                "chr7_agreeing_bins": chr7.agreeing_bins if chr7 else None,
                "chr8_cn": chr8.rounded_copy_number if chr8 else None,
                "chr8_agreeing_bins": chr8.agreeing_bins if chr8 else None,
            },
        )
    )
    return checks


def _verify_release_checksums(envelope_root: Path) -> tuple[bool, int]:
    manifest = envelope_root / "release" / "checksums.sha256"
    if not manifest.is_file() or manifest.stat().st_size == 0:
        return False, 0
    envelope_root = envelope_root.resolve()
    verified = 0
    for raw in manifest.read_text(encoding="utf-8").splitlines():
        if not raw.strip():
            continue
        try:
            expected, relative = raw.split(maxsplit=1)
        except ValueError:
            return False, verified
        target = (envelope_root / relative.strip()).resolve()
        if not target.is_relative_to(envelope_root) or not target.is_file():
            return False, verified
        if sha256_file(target) != expected:
            return False, verified
        verified += 1
    return verified > 0, verified


def run_system_smoke(
    output_dir: Path,
    qc_policy: QCPolicy,
    sniffles_policy: SnifflesPolicy,
    cnv_policy: QDNAseqPolicy,
    *,
    qdnaseq_script: Path,
    runner: StreamingCommandRunner | None = None,
    samtools: str = "samtools",
    cramino: str = "cramino",
    sniffles: str = "sniffles",
    rscript: str = "Rscript",
    threads: int = 2,
    pipeline_version: str = "UNKNOWN",
    git_commit: str = "SYSTEM_SMOKE",
) -> SystemSmokeReport:
    """Exercise installed local tools plus the canonical QDNAseq+ACE run path."""

    if threads < 1:
        raise ValueError("threads must be at least 1")
    if not qdnaseq_script.is_file():
        raise ValueError(f"QDNAseq runner script does not exist: {qdnaseq_script}")

    output_dir = output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    summary_path = output_dir / "system-smoke.report.json"
    if summary_path.exists():
        raise ValueError(f"Refusing to overwrite existing system-smoke report: {summary_path}")

    command_runner = runner or SubprocessRunner()
    local_report = run_local_smoke(
        output_dir / "sv-qc",
        qc_policy,
        sniffles_policy,
        runner=command_runner,
        samtools=samtools,
        cramino=cramino,
        sniffles=sniffles,
        threads=threads,
        pipeline_version=pipeline_version,
        git_commit=git_commit,
    )

    manifest, reference_lock = _build_cnv_fixture(
        output_dir / "cnv-fixture",
        runner=command_runner,
        samtools=samtools,
        threads=threads,
    )
    register_qdnaseq_extension(
        QDNAseqExtensionSettings(
            policy=cnv_policy,
            rscript=rscript,
            script=qdnaseq_script.resolve(),
        )
    )

    output_base = output_dir / "runs"
    config = RunConfiguration(
        manifest=manifest,
        reference_lock=reference_lock,
        output_base=output_base,
        run_id=SYSTEM_SMOKE_RUN_ID,
        pipeline_version=pipeline_version,
        git_commit=git_commit,
        qc_policy=qc_policy,
        sniffles_policy=sniffles_policy,
        threads=threads,
        executables={
            "samtools": samtools,
            "cramino": cramino,
            "sniffles": sniffles,
            "minimap2": "minimap2",
            "dorado": "dorado",
        },
    )
    first_report, first_release = run_pipeline(config, runner=command_runner)
    if not first_report.passed or first_release is None:
        raise ValueError("Canonical CNV system-smoke run failed: " + first_report.verdict_reason)

    envelope_root = output_base / SYSTEM_SMOKE_RUN_ID / SYSTEM_SMOKE_SAMPLE_ID
    qdnaseq_path = envelope_root / "evidence" / "cnv" / f"{SYSTEM_SMOKE_SAMPLE_ID}.qdnaseq.json"
    result_path = envelope_root / "normalized" / f"{SYSTEM_SMOKE_SAMPLE_ID}.result.json"
    html_path = envelope_root / "reports" / f"{SYSTEM_SMOKE_SAMPLE_ID}.report.html"
    workbook_path = envelope_root / "reports" / f"{SYSTEM_SMOKE_SAMPLE_ID}.results.xlsx"
    run_report_path = envelope_root / "provenance" / "run.json"
    release_path = envelope_root / "release" / "release.json"
    checksums_path = envelope_root / "release" / "checksums.sha256"

    required_outputs = (
        qdnaseq_path,
        result_path,
        html_path,
        workbook_path,
        run_report_path,
        release_path,
        checksums_path,
    )
    missing_outputs = [
        str(path.relative_to(envelope_root))
        for path in required_outputs
        if not path.is_file() or path.stat().st_size == 0
    ]
    if missing_outputs:
        raise ValueError(
            "Canonical system-smoke completed without required outputs: "
            + ", ".join(missing_outputs)
        )

    qdnaseq_report = QDNAseqCallReport.model_validate_json(qdnaseq_path.read_text(encoding="utf-8"))
    normalized = PipelineResult.model_validate_json(result_path.read_text(encoding="utf-8"))
    first_cnv_stage = next(
        (item for item in first_report.stages if item.stage == StageId.CNV),
        None,
    )

    checks: list[ValidationCheck] = [
        ValidationCheck(
            name="sv_qc_local_smoke",
            status=CheckStatus.PASS if local_report.verdict == Verdict.PASS else CheckStatus.FAIL,
            message=(
                "Synthetic long-read intake, Cramino, Sniffles2 and basic report rendering passed."
            ),
            details={"accepted_sv_candidates": local_report.sniffles.accepted_record_count},
        ),
        ValidationCheck(
            name="canonical_cnv_pipeline",
            status=(
                CheckStatus.PASS
                if first_cnv_stage is not None
                and first_cnv_stage.status == ModuleRunStatus.COMPLETED
                else CheckStatus.FAIL
            ),
            message=(
                "The canonical run envelope completed the live QDNAseq+ACE CNV stage."
                if first_cnv_stage is not None
                and first_cnv_stage.status == ModuleRunStatus.COMPLETED
                else "The canonical run did not complete the live QDNAseq+ACE CNV stage."
            ),
            details={"stage_count": len(first_report.stages)},
        ),
    ]
    checks.extend(cnv_truth_checks(qdnaseq_report, cnv_policy))

    cnv_module = next(
        (item for item in normalized.modules if item.module == AnalysisModule.CNV),
        None,
    )
    html_has_cnv = "Copy-number analysis — QDNAseq + ACE" in html_path.read_text(encoding="utf-8")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    required_sheets = {"CNV Fits", "CNV Consensus", "CNV Segments"}
    workbook_has_cnv = required_sheets.issubset(set(workbook.sheetnames))
    workbook.close()
    tool_names = {item.name for item in normalized.provenance.tools}
    report_bundle_ok = (
        cnv_module is not None
        and cnv_module.status == ModuleRunStatus.COMPLETED
        and {"QDNAseq", "ACE"}.issubset(tool_names)
        and html_has_cnv
        and workbook_has_cnv
    )
    checks.append(
        ValidationCheck(
            name="integrated_cnv_reporting",
            status=CheckStatus.PASS if report_bundle_ok else CheckStatus.FAIL,
            message=(
                "Normalized JSON, provenance, HTML and Excel all contain the integrated CNV result."
                if report_bundle_ok
                else (
                    "The integrated CNV result is incomplete across JSON, provenance, "
                    "HTML or Excel."
                )
            ),
            details={
                "html_has_cnv": html_has_cnv,
                "workbook_has_cnv_sheets": workbook_has_cnv,
                "cnv_module_completed": bool(
                    cnv_module is not None and cnv_module.status == ModuleRunStatus.COMPLETED
                ),
                "qdnaseq_in_provenance": "QDNAseq" in tool_names,
                "ace_in_provenance": "ACE" in tool_names,
            },
        )
    )

    checksum_ok, checksum_count = _verify_release_checksums(envelope_root)
    checks.append(
        ValidationCheck(
            name="release_checksums",
            status=CheckStatus.PASS if checksum_ok else CheckStatus.FAIL,
            message=(
                "Every checksummed release artifact independently matched SHA256."
                if checksum_ok
                else "Release checksum verification failed."
            ),
            details={"verified_artifacts": checksum_count},
        )
    )

    resumed_report, resumed_release = run_pipeline(config, runner=command_runner)
    resumed_cnv_stage = next(
        (item for item in resumed_report.stages if item.stage == StageId.CNV),
        None,
    )
    resume_ok = (
        resumed_report.passed
        and resumed_release is not None
        and resumed_cnv_stage is not None
        and resumed_cnv_stage.resumed
    )
    checks.append(
        ValidationCheck(
            name="content_addressed_cnv_resume",
            status=CheckStatus.PASS if resume_ok else CheckStatus.FAIL,
            message=(
                "An unchanged second run resumed the verified CNV stage instead of recomputing it."
                if resume_ok
                else "The unchanged CNV stage did not resume as expected."
            ),
            details={"cnv_resumed": bool(resumed_cnv_stage and resumed_cnv_stage.resumed)},
        )
    )

    checksum_ok_after_resume, checksum_count_after_resume = _verify_release_checksums(envelope_root)
    checks.append(
        ValidationCheck(
            name="release_checksums_after_resume",
            status=(CheckStatus.PASS if checksum_ok_after_resume else CheckStatus.FAIL),
            message=(
                "Release checksums remained internally valid after the resumed run."
                if checksum_ok_after_resume
                else "Release checksums failed after the resumed run."
            ),
            details={"verified_artifacts": checksum_count_after_resume},
        )
    )

    verdict = Verdict.FAIL if any(item.status == CheckStatus.FAIL for item in checks) else Verdict.PASS
    report = SystemSmokeReport(
        verdict=verdict,
        checks=checks,
        output_paths={
            "local_smoke": str(output_dir / "sv-qc" / "local-smoke.report.json"),
            "cnv_qdnaseq": str(qdnaseq_path),
            "normalized_result": str(result_path),
            "html_report": str(html_path),
            "excel_report": str(workbook_path),
            "run_report": str(run_report_path),
            "release": str(release_path),
            "release_checksums": str(checksums_path),
        },
        limitations=[
            (
                "The fixtures are deterministic synthetic engineering data, not "
                "biological validation material."
            ),
            ("The CNV fixture deliberately targets the packaged GRCh37/QDNAseq.hg19 lane."),
            (
                "A passing system smoke proves local execution, integration, packaging "
                "and resume behavior for the exercised paths."
            ),
        ],
    )
    write_json(report, summary_path)
    if verdict == Verdict.FAIL:
        failures = [item.name for item in checks if item.status == CheckStatus.FAIL]
        raise ValueError("System smoke failed: " + ", ".join(failures))
    return report
