from __future__ import annotations

import base64
import html
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from ..models import (
    AnalysisModule,
    ModuleOutcome,
    ModuleRunStatus,
    PipelineResult,
    Provenance,
)
from ..pipeline import runner as pipeline_runner
from ..pipeline.envelope import Artifact, sha256_file
from ..pipeline.runner import StageImplementation, StagePlan, StageResult
from ..pipeline.stages import StageId, VerificationStatus
from .qdnaseq import QDNAseqCallReport, QDNAseqPolicy, run_qdnaseq_ace

#: Names this extension in the result's assemble reason and in both contribution slots.
EXTENSION_ID = "qdnaseq-ace-v1"

CNV_DIR = "evidence/cnv/qdnaseq"
CNV_REPORT = "evidence/cnv/{sample}.qdnaseq.json"


@dataclass(frozen=True)
class QDNAseqExtensionSettings:
    policy: QDNAseqPolicy
    rscript: str = "Rscript"
    script: Path = Path("scripts/run_qdnaseq_ace.R")


_SETTINGS: QDNAseqExtensionSettings | None = None


def _settings() -> QDNAseqExtensionSettings:
    if _SETTINGS is None:
        raise RuntimeError("QDNAseq extension has not been registered")
    return _SETTINGS


def _requested(ctx: pipeline_runner.RunContext) -> bool:
    return AnalysisModule.CNV in set(ctx.manifest.analysis.modules)


def _probe_r_packages(ctx: pipeline_runner.RunContext) -> dict[str, str]:
    settings = _settings()
    r_version = ctx.runner.run([settings.rscript, "--version"], timeout_seconds=60)
    if r_version.returncode != 0:
        raise ValueError(f"Rscript version probe failed: {r_version.stderr.strip()}")
    expression = (
        "cat(as.character(packageVersion('QDNAseq')), '\\n');"
        "cat(as.character(packageVersion('ACE')), '\\n');"
        "cat(as.character(packageVersion('DNAcopy')), '\\n')"
    )
    packages = ctx.runner.run([settings.rscript, "-e", expression], timeout_seconds=60)
    if packages.returncode != 0:
        diagnostic = packages.stderr.strip()[-2000:]
        raise ValueError(f"QDNAseq/ACE R packages are unavailable: {diagnostic}")
    lines = [line.strip() for line in packages.stdout.splitlines() if line.strip()]
    if len(lines) < 3:
        raise ValueError("could not determine QDNAseq, ACE and DNAcopy versions")
    r_text = r_version.stderr.strip() or r_version.stdout.strip()
    return {
        "Rscript": r_text,
        "QDNAseq": lines[0],
        "ACE": lines[1],
        "DNAcopy": lines[2],
    }


def _cnv_plan(ctx: pipeline_runner.RunContext) -> StagePlan:
    if not _requested(ctx):
        return StagePlan(parameters={"requested": False}, tool_versions={})
    settings = _settings()
    if not settings.script.is_file():
        raise ValueError(f"QDNAseq R runner not found: {settings.script}")
    versions = _probe_r_packages(ctx)
    bam = Path(ctx.manifest.input.path)
    return StagePlan(
        parameters={
            "requested": True,
            "profile": settings.policy.profile_id,
            "bin_sizes_kbp": settings.policy.bin_sizes_kbp,
            "primary_bin_size_kbp": settings.policy.primary_bin_size_kbp,
            "ace_penalty": settings.policy.ace_penalty,
            "ploidy_min": settings.policy.ploidy_min,
            "ploidy_max": settings.policy.ploidy_max,
            "ploidy_step": settings.policy.ploidy_step,
            "threads": ctx.config.threads,
        },
        tool_versions=versions,
        external_inputs=(
            (bam.name, sha256_file(bam)),
            (settings.script.name, sha256_file(settings.script)),
        ),
    )


def _record_cnv_outputs(
    ctx: pipeline_runner.RunContext,
    report: QDNAseqCallReport,
) -> list[Artifact]:
    artifacts: list[Artifact] = []
    relative_report = ctx.path(CNV_REPORT)
    artifacts.append(
        ctx.envelope.atomic_write_text(relative_report, report.model_dump_json(indent=2) + "\n")
    )
    for name in report.output_files:
        if name.lower().endswith(".rds"):
            continue
        artifacts.append(ctx.envelope.fingerprint(f"{CNV_DIR}/{name}"))
    return artifacts


def _cnv_execute(ctx: pipeline_runner.RunContext, plan: StagePlan) -> StageResult:
    del plan
    if not _requested(ctx):
        return StageResult(
            status=ModuleRunStatus.NOT_RUN,
            reason="CNV was not requested in the sample manifest.",
        )
    settings = _settings()
    report = run_qdnaseq_ace(
        bam=Path(ctx.manifest.input.path),
        sample_id=ctx.sample_id,
        genome_build=ctx.manifest.assay.genome_build,
        reference_lock=ctx.config.reference_lock,
        policy=settings.policy,
        output_dir=ctx.envelope.path(CNV_DIR),
        script=settings.script,
        runner=ctx.runner,
        rscript=settings.rscript,
        threads=ctx.config.threads,
    )
    artifacts = _record_cnv_outputs(ctx, report)
    bins = ", ".join(str(value) for value in settings.policy.bin_sizes_kbp)
    reason = (
        f"QDNAseq+ACE completed at {bins} kbp; primary "
        f"{report.primary_fit.bin_size_kbp} kbp fit: cellularity "
        f"{report.primary_fit.cellularity:.3f}, ploidy {report.primary_fit.ploidy:.3f}; "
        f"{len(report.events)} normalized CNV event(s)."
    )
    return StageResult(
        status=report.status,
        reason=reason,
        outputs=artifacts,
        tools=report.tools,
        warnings=report.warnings,
        limitations=report.limitations,
    )


def _load_cnv(ctx: pipeline_runner.RunContext) -> QDNAseqCallReport | None:
    path = ctx.envelope.path(ctx.path(CNV_REPORT))
    if not path.is_file():
        return None
    return QDNAseqCallReport.model_validate_json(path.read_text(encoding="utf-8"))


def _replace_module(
    modules: Sequence[ModuleOutcome],
    replacement: ModuleOutcome,
) -> list[ModuleOutcome]:
    result = [item for item in modules if item.module != replacement.module]
    result.append(replacement)
    return sorted(result, key=lambda item: item.module.value)


def _merge_provenance(base: Provenance, cnv: QDNAseqCallReport) -> Provenance:
    return base.model_copy(update={"tools": [*base.tools, *cnv.tools]})


def _cnv_html_section(ctx: pipeline_runner.RunContext, cnv: QDNAseqCallReport) -> str:
    fit_rows = "".join(
        "<tr>"
        f"<td>{fit.bin_size_kbp}</td><td>{fit.cellularity:.3f}</td>"
        f"<td>{fit.ploidy:.3f}</td><td>{fit.fit_error:.6g}</td>"
        f"<td>{fit.segment_count}</td></tr>"
        for fit in sorted(cnv.fits, key=lambda value: value.bin_size_kbp)
    )
    chromosome_rows = "".join(
        "<tr>"
        f"<td>{html.escape(chromosome.chromosome)}</td>"
        f"<td>{chromosome.median_copy_number:.3f}</td>"
        f"<td>{chromosome.rounded_copy_number}</td>"
        f"<td>{chromosome.agreeing_bins}/{chromosome.contributing_bins}</td>"
        f"<td>{chromosome.min_copy_number:.3f}–{chromosome.max_copy_number:.3f}</td>"
        "</tr>"
        for chromosome in cnv.chromosome_consensus
    )
    images: list[str] = []
    for label, name in (
        ("ACE purity/ploidy fit landscape", cnv.primary_fit.fit_plot),
        ("Absolute copy-number profile", cnv.primary_fit.copy_number_plot),
    ):
        path = ctx.envelope.path(f"{CNV_DIR}/{name}")
        if not path.is_file():
            continue
        encoded = base64.b64encode(path.read_bytes()).decode("ascii")
        images.append(
            f"<h3>{html.escape(label)}</h3>"
            f"<img alt='{html.escape(label)}' style='max-width:100%;height:auto' "
            f"src='data:image/png;base64,{encoded}'>"
        )
    return (
        "<section><h2>Copy-number analysis — QDNAseq + ACE</h2>"
        f"<p><strong>Primary:</strong> {cnv.primary_fit.bin_size_kbp} kbp; "
        f"cellularity {cnv.primary_fit.cellularity:.3f}; "
        f"ploidy {cnv.primary_fit.ploidy:.3f}; "
        f"fit error {cnv.primary_fit.fit_error:.6g}.</p>"
        "<h3>Multi-resolution fits</h3><table><thead><tr><th>Bin (kbp)</th>"
        "<th>Cellularity</th><th>Ploidy</th><th>Fit error</th><th>Segments</th>"
        f"</tr></thead><tbody>{fit_rows}</tbody></table>"
        "<h3>Chromosome-level consensus</h3><table><thead><tr><th>Chromosome</th>"
        "<th>Median CN</th><th>Rounded CN</th><th>Agreement</th><th>Range</th>"
        f"</tr></thead><tbody>{chromosome_rows}</tbody></table>" + "".join(images) + "</section>"
    )


def _enrich_workbook(
    path: Path,
    ctx: pipeline_runner.RunContext,
    cnv: QDNAseqCallReport,
) -> None:
    workbook = load_workbook(path)
    for title in ("CNV Fits", "CNV Consensus", "CNV Segments"):
        if title in workbook.sheetnames:
            del workbook[title]

    fits_sheet = workbook.create_sheet("CNV Fits")
    fits_sheet.append(
        [
            "bin_size_kbp",
            "cellularity",
            "ploidy",
            "fit_error",
            "candidate_count",
            "segment_count",
        ]
    )
    for fit in sorted(cnv.fits, key=lambda value: value.bin_size_kbp):
        fits_sheet.append(
            [
                fit.bin_size_kbp,
                fit.cellularity,
                fit.ploidy,
                fit.fit_error,
                fit.candidate_count,
                fit.segment_count,
            ]
        )

    consensus_sheet = workbook.create_sheet("CNV Consensus")
    consensus_sheet.append(
        [
            "chromosome",
            "median_copy_number",
            "rounded_copy_number",
            "agreeing_bins",
            "contributing_bins",
            "min_copy_number",
            "max_copy_number",
        ]
    )
    for chromosome in cnv.chromosome_consensus:
        consensus_sheet.append(
            [
                chromosome.chromosome,
                chromosome.median_copy_number,
                chromosome.rounded_copy_number,
                chromosome.agreeing_bins,
                chromosome.contributing_bins,
                chromosome.min_copy_number,
                chromosome.max_copy_number,
            ]
        )

    segment_sheet = workbook.create_sheet("CNV Segments")
    segment_path = ctx.envelope.path(f"{CNV_DIR}/{cnv.primary_fit.segment_file}")
    with segment_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            segment_sheet.append(line.rstrip("\n").split("\t"))
    workbook.save(path)


def _enrich_result(ctx: pipeline_runner.RunContext, result: PipelineResult) -> PipelineResult:
    """Fold this run's CNV evidence into the result the runner already assembled.

    Additive by construction: it receives the finished result and returns one with the CNV
    events, module outcome, provenance and warnings added. It cannot drop what it does not
    know about, which is what the previous whole-stage replacement did to the SV consensus.
    """
    cnv = _load_cnv(ctx)
    if cnv is None:
        return result
    outcome = ModuleOutcome(
        module=AnalysisModule.CNV,
        status=cnv.status,
        reason=(
            f"QDNAseq+ACE multi-bin CNV completed; primary "
            f"{cnv.primary_fit.bin_size_kbp} kbp, cellularity "
            f"{cnv.primary_fit.cellularity:.3f}, ploidy {cnv.primary_fit.ploidy:.3f}."
        ),
        tools=cnv.tools,
    )
    return result.model_copy(
        update={
            "events": [*cnv.events, *result.events],
            "modules": _replace_module(result.modules, outcome),
            "provenance": _merge_provenance(result.provenance, cnv),
            "warnings": [*result.warnings, *cnv.warnings, *cnv.limitations],
        }
    )


def _enrich_reports(ctx: pipeline_runner.RunContext, html_path: Path, xlsx_path: Path) -> bool:
    """Add the CNV section to the already-rendered reviewer artifacts.

    Returns whether anything was written. A run that did not request CNV, or whose CNV stage
    recorded NOT_RUN or FAILED, leaves no report to fold in — and must not be described as
    carrying a CNV section it does not have.
    """
    cnv = _load_cnv(ctx)
    if cnv is None:
        return False
    document = html_path.read_text(encoding="utf-8")
    section = _cnv_html_section(ctx, cnv)
    marker = "<section><h2>Warnings and limitations</h2>"
    if marker in document:
        document = document.replace(marker, section + marker, 1)
    else:
        document = document.replace("</main>", section + "</main>", 1)
    html_path.write_text(document, encoding="utf-8")
    _enrich_workbook(xlsx_path, ctx, cnv)
    return True


def register_qdnaseq_extension(settings: QDNAseqExtensionSettings) -> None:
    """Install the live CNV lane for this process.

    Three things this deliberately no longer does. It does not rewrite ``SPEC_BY_STAGE``:
    the graph is data, built once at import, and a registration that edited it made the
    verification status depend on registration order — preflight, which does not register,
    reported CNV as having no adapter for runs that then executed a real QDNAseq analysis.
    It does not replace the assemble stage, and it does not replace the report stage. It
    supplies the CNV stage, which is its own, and contributes to the other two.
    """
    global _SETTINGS
    _SETTINGS = settings
    pipeline_runner.IMPLEMENTATIONS[StageId.CNV] = StageImplementation(
        _cnv_plan,
        _cnv_execute,
        # The declared graph calls CNV not_implemented because nothing is wired in by
        # default. This adapter is exercised against real QDNAseq and ACE in CI, and says
        # so here rather than by editing the graph.
        verification=VerificationStatus.VERIFIED_WITH_REAL_TOOL,
    )
    pipeline_runner.register_result_contribution(
        pipeline_runner.ResultContribution(
            extension_id=EXTENSION_ID,
            source_artifacts=(CNV_REPORT,),
            enrich=_enrich_result,
        )
    )
    pipeline_runner.register_report_contribution(
        pipeline_runner.ReportContribution(
            extension_id=EXTENSION_ID,
            source_artifacts=(CNV_REPORT,),
            enrich=_enrich_reports,
        )
    )
