from __future__ import annotations

import json
from collections.abc import Iterable
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import GenomicEvent, PipelineResult, ResolvedResourceContext
from .reporting import (
    caller_count,
    fusion_assessment,
    fusion_review_events,
    gene_pair_label,
    is_structural_variant,
    key_findings,
    maximum_support,
    pathology_label,
    release_state,
    review_priority,
)
from .target_coverage import TargetCoverageReport

HEADER_FILL = PatternFill("solid", fgColor="0B557A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WARNING_FILL = PatternFill("solid", fgColor="FFF7ED")
CRITICAL_FILL = PatternFill("solid", fgColor="FFF1F2")


def _style_sheet(sheet: Worksheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 58)
        column_index = column[0].column
        if column_index is not None:
            sheet.column_dimensions[get_column_letter(column_index)].width = max(width, 11)


def _write_table(sheet: Worksheet, headers: list[str], rows: Iterable[list[object]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    _style_sheet(sheet)


def _apply_row_fill(sheet: Worksheet, column_name: str) -> None:
    headers = [cell.value for cell in sheet[1]]
    if column_name not in headers:
        return
    column = headers.index(column_name) + 1
    fills = {"HEMATOLOGY_REVIEW": CRITICAL_FILL, "FUSION_REVIEW": WARNING_FILL}
    for row_index in range(2, sheet.max_row + 1):
        fill = fills.get(str(sheet.cell(row_index, column).value or ""))
        if fill is not None:
            for cell in sheet[row_index]:
                cell.fill = fill


def _locus(event: GenomicEvent, *, secondary: bool = False) -> str:
    locus = event.secondary if secondary else event.primary
    if locus is None:
        return ""
    return f"{locus.chromosome}:{locus.start}-{locus.end}"


def _cytobands(event: GenomicEvent) -> str:
    bands = [event.primary.cytoband_start or ""]
    if event.secondary is not None:
        bands.append(event.secondary.cytoband_start or "")
    return " ↔ ".join(item for item in bands if item)


def _support(event: GenomicEvent) -> str:
    return ", ".join(
        f"{item.caller}={item.support_reads if item.support_reads is not None else 'n/a'}"
        for item in event.evidence
    )


def _coverage(event: GenomicEvent) -> str:
    return " / ".join(
        "n/a" if depth is None else f"{depth:.1f}x" for depth in event.breakpoint_mean_depths
    )


def _preferred_transcript(event: GenomicEvent, index: int) -> str:
    if index >= len(event.breakpoint_annotations):
        return ""
    annotation = event.breakpoint_annotations[index]
    preferred = next((item for item in annotation.transcripts if item.preferred), None)
    if preferred is None and annotation.transcripts:
        preferred = annotation.transcripts[0]
    if preferred is None:
        return ""
    location: str = preferred.region
    if preferred.exon_number is not None:
        location += f" {preferred.exon_number}"
    elif preferred.intron_number is not None:
        location += f" {preferred.intron_number}"
    return f"{preferred.gene_name}/{preferred.transcript_id} ({location})"


def _resource_summary(result: PipelineResult) -> list[list[object]]:
    context = result.reference_context
    if not isinstance(context, ResolvedResourceContext):
        return [["Reference context", "legacy_unspecified"]]
    releases = context.resource_releases
    return [
        ["Genome assembly", releases.get("reference.genome_fasta", context.genome_build.value)],
        ["ReferenceBundle", f"{context.reference_bundle_id} ({context.reference_bundle_version})"],
        ["BAM dictionary contract", context.reference_dictionary_contract.value],
        ["GENCODE", releases.get("reference.gencode_gtf", "unspecified")],
        ["MANE", releases.get("reference.mane_gff3", "unspecified")],
        ["Cytobands", releases.get("reference.cytobands", "unspecified")],
        [
            "PanelBundle",
            (
                f"{context.panel_bundle_id} ({context.panel_bundle_version})"
                if context.panel_bundle_id is not None
                else "NOT_APPLICABLE"
            ),
        ],
        ["KnowledgeBundle", f"{context.knowledge_bundle_id} ({context.knowledge_bundle_version})"],
    ]


def _write_summary(workbook: Workbook, result: PipelineResult) -> None:
    structural = [event for event in result.events if is_structural_variant(event)]
    sheet = workbook.create_sheet("00_Summary")
    _write_table(
        sheet,
        ["Field", "Value"],
        [
            ["Sample ID", result.manifest.sample_id],
            ["Run ID", result.manifest.run_id],
            ["Assay", result.manifest.assay.mode.value],
            ["Genome build", result.manifest.assay.genome_build.value],
            ["QC verdict", result.qc.verdict.value],
            ["Pipeline release status", result.release_status.value],
            ["Key findings for review", len(key_findings(structural))],
            ["All normalized structural variants", len(structural)],
            [
                "Release-state interpretation",
                "BENCHMARK_REQUIRED means reviewable evidence without an assay-specific analytical "
                "release gate; it does not mean irrelevant, negative, or absent.",
            ],
            [
                "Pathology-association interpretation",
                "Pathology names and DOIDs describe source-database associations for a matched "
                "gene pair; they are not diagnoses for this sample.",
            ],
            ["ISCN proposal", result.iscn.notation],
            ["ISCN profile", result.iscn.conformance_profile],
            *_resource_summary(result),
        ],
    )
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 1).font = Font(bold=True, color="334155")
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 90


def _write_key_findings(workbook: Workbook, events: list[GenomicEvent]) -> None:
    sheet = workbook.create_sheet("01_Key_Findings")
    _write_table(
        sheet,
        [
            "review_priority",
            "gene_pair",
            "event_id",
            "type",
            "locus_1",
            "locus_2",
            "cytobands",
            "technical_confidence",
            "caller_count",
            "max_support_reads",
            "caller_support",
            "breakpoint_coverage",
            "fusion_assessment",
            "knowledge_match",
            "associated_pathologies",
            "release_state",
        ],
        [
            [
                review_priority(event),
                gene_pair_label(event),
                event.event_id,
                event.event_type.value,
                _locus(event),
                _locus(event, secondary=True),
                _cytobands(event),
                event.confidence,
                caller_count(event),
                maximum_support(event),
                _support(event),
                _coverage(event),
                fusion_assessment(event),
                event.known_rearrangement or "",
                pathology_label(event),
                release_state(event),
            ]
            for event in key_findings(events)
        ],
    )
    _apply_row_fill(sheet, "review_priority")


def _write_coverage(
    workbook: Workbook,
    target: TargetCoverageReport | None,
    selection: TargetCoverageReport | None,
) -> None:
    sheet = workbook.create_sheet("12_AS_Coverage")
    if target is None:
        _write_table(
            sheet,
            ["Status", "Explanation"],
            [["NOT_ASSESSED", "No target-coverage sidecar was supplied to the renderer."]],
        )
        return
    thresholds = [f"{value}x" for value in target.policy.thresholds]
    _write_table(
        sheet,
        [
            "region_id",
            "chromosome",
            "start",
            "end",
            "length_bp",
            "mean_depth_x",
            *[f"bases_at_{threshold}_fraction" for threshold in thresholds],
        ],
        [
            [
                region.region_id,
                region.chromosome,
                region.start,
                region.end,
                region.end - region.start,
                region.mean_depth,
                *[region.fraction_at_threshold[threshold] for threshold in thresholds],
            ]
            for region in sorted(target.regions, key=lambda item: item.mean_depth)
        ],
    )
    for column in range(7, 7 + len(thresholds)):
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, column).number_format = "0.0%"
    sheet.insert_rows(1, 5)
    sheet["A1"] = "Adaptive-sampling analysis ROI coverage"
    sheet["A1"].font = Font(size=16, bold=True, color="0B557A")
    summary = [
        ("Targets assessed", target.summary_metrics.get("region_count")),
        (
            "Interval-weighted mean depth",
            target.summary_metrics.get("interval_weighted_mean_depth"),
        ),
        ("Minimum target mean depth", target.summary_metrics.get("minimum_region_mean_depth")),
        (
            "Buffered selection mean depth",
            selection.summary_metrics.get("interval_weighted_mean_depth") if selection else "n/a",
        ),
    ]
    for row, (label, value) in enumerate(summary, start=2):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
    sheet.freeze_panes = "A7"
    sheet.auto_filter.ref = f"A6:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    for cell in sheet[6]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _write_qc(workbook: Workbook, result: PipelineResult) -> None:
    sheet = workbook.create_sheet("01_QC")
    _write_table(sheet, ["Metric", "Value"], list(map(list, result.qc.metrics.items())))


def _write_cnv(workbook: Workbook, result: PipelineResult) -> None:
    chromosomes = workbook.create_sheet("02_CNV_Chromosomes")
    chromosome_events = [
        event for event in result.events if "chromosome_" in event.event_type.value
    ]
    _write_table(
        chromosomes,
        ["event_id", "event_type", "chromosome", "copy_number", "confidence", "release_state"],
        [
            [
                event.event_id,
                event.event_type.value,
                event.primary.chromosome,
                event.copy_number,
                event.confidence,
                release_state(event),
            ]
            for event in chromosome_events
        ],
    )
    segments = workbook.create_sheet("03_CNV_Segments")
    segment_events = [
        event
        for event in result.events
        if event.event_type.value in {"deletion", "duplication"} and event.copy_number is not None
    ]
    _write_table(
        segments,
        [
            "event_id",
            "type",
            "chromosome",
            "start",
            "end",
            "band_start",
            "band_end",
            "copy_number",
            "confidence",
            "release_state",
        ],
        [
            [
                event.event_id,
                event.event_type.value,
                event.primary.chromosome,
                event.primary.start,
                event.primary.end,
                event.primary.cytoband_start,
                event.primary.cytoband_end,
                event.copy_number,
                event.confidence,
                release_state(event),
            ]
            for event in segment_events
        ],
    )


def _write_sv(workbook: Workbook, events: list[GenomicEvent]) -> None:
    sheet = workbook.create_sheet("04_SV")
    _write_table(
        sheet,
        [
            "event_id",
            "review_priority",
            "gene_pair",
            "type",
            "length_bp",
            "locus_1",
            "locus_2",
            "cytobands",
            "caller_count",
            "caller_support",
            "technical_confidence",
            "breakpoint_coverage",
            "as_observability",
            "technical_flags",
            "aml_relevance",
            "known_rearrangement",
            "associated_pathologies",
            "fusion_assessment",
            "validation_status",
            "release_state",
            "source_event_ids",
            "evidence_json",
            "reportable",
        ],
        [
            [
                event.event_id,
                review_priority(event),
                gene_pair_label(event),
                event.event_type.value,
                event.length_bp,
                _locus(event),
                _locus(event, secondary=True),
                _cytobands(event),
                caller_count(event),
                _support(event),
                event.confidence,
                _coverage(event),
                event.observability.value,
                ", ".join(event.technical_flags),
                event.aml_relevance or "",
                event.known_rearrangement or "",
                pathology_label(event),
                fusion_assessment(event),
                event.validation_status.value,
                release_state(event),
                ", ".join(event.source_event_ids),
                json.dumps([item.model_dump(mode="json") for item in event.evidence]),
                event.reportable,
            ]
            for event in events
        ],
    )
    _apply_row_fill(sheet, "review_priority")


def _write_fusions(workbook: Workbook, events: list[GenomicEvent]) -> None:
    sheet = workbook.create_sheet("05_Fusions")
    _write_table(
        sheet,
        [
            "review_priority",
            "gene_pair",
            "event_id",
            "type",
            "breakpoint_1",
            "breakpoint_2",
            "cytobands",
            "preferred_context_1",
            "preferred_context_2",
            "orientation",
            "frame_status",
            "caller_support",
            "breakpoint_coverage",
            "technical_confidence",
            "knowledge_match",
            "associated_pathologies",
            "fusion_assessment",
            "release_state",
        ],
        [
            [
                review_priority(event),
                gene_pair_label(event),
                event.event_id,
                event.event_type.value,
                _locus(event),
                _locus(event, secondary=True),
                _cytobands(event),
                _preferred_transcript(event, 0),
                _preferred_transcript(event, 1),
                (
                    event.fusion_evidence.orientation or "unknown"
                    if event.fusion_evidence is not None
                    else "unknown"
                ),
                (
                    event.fusion_evidence.frame_status
                    if event.fusion_evidence is not None
                    else "unknown"
                ),
                _support(event),
                _coverage(event),
                event.confidence,
                event.known_rearrangement or "",
                pathology_label(event),
                fusion_assessment(event),
                release_state(event),
            ]
            for event in fusion_review_events(events)
        ],
    )
    _apply_row_fill(sheet, "review_priority")


def _write_supporting_sheets(workbook: Workbook, result: PipelineResult) -> None:
    iscn = workbook.create_sheet("06_ISCN")
    _write_table(
        iscn,
        ["Field", "Value"],
        [
            ["Notation", result.iscn.notation],
            ["Edition", result.iscn.standard_edition],
            ["Conformance profile", result.iscn.conformance_profile],
            ["Review status", result.iscn.review_status.value],
            ["Source event IDs", ", ".join(result.iscn.source_event_ids)],
        ],
    )
    warnings = workbook.create_sheet("07_Warnings")
    warning_rows: list[list[object]] = [
        [item] for item in result.warnings + result.qc.warnings + result.iscn.warnings
    ]
    _write_table(warnings, ["Warning"], warning_rows)
    for cell in warnings["A"][1:]:
        cell.fill = WARNING_FILL

    methods = workbook.create_sheet("08_Methods_Versions")
    context = result.reference_context
    method_rows: list[list[object]] = [
        [
            tool.name,
            tool.version,
            json.dumps(tool.parameters, sort_keys=True),
            tool.container_digest,
        ]
        for tool in result.provenance.tools
    ]
    if isinstance(context, ResolvedResourceContext):
        method_rows.extend(
            [
                f"resource:{name}",
                context.resource_releases.get(name, "unspecified"),
                "",
                checksum,
            ]
            for name, checksum in sorted(context.resource_checksums.items())
        )
    _write_table(
        methods,
        ["Tool", "Version", "Parameters", "Digest / SHA256"],
        method_rows,
    )

    run_log = workbook.create_sheet("09_Run_Log")
    _write_table(
        run_log,
        ["Field", "Value"],
        [
            ["Pipeline version", result.provenance.pipeline_version],
            ["Git commit", result.provenance.git_commit],
            ["Created at", result.provenance.created_at.isoformat()],
            [
                "Reference checksums",
                json.dumps(result.provenance.reference_checksums, sort_keys=True),
            ],
        ],
    )
    modules = workbook.create_sheet("10_Module_Status")
    _write_table(
        modules,
        ["Module", "Status", "Reason", "Tools"],
        [
            [
                outcome.module.value,
                outcome.status.value,
                outcome.reason,
                ", ".join(f"{tool.name} {tool.version}" for tool in outcome.tools),
            ]
            for outcome in result.modules
        ],
    )


def render_workbook(
    result: PipelineResult,
    output_path: Path,
    *,
    target_coverage: TargetCoverageReport | None = None,
    selection_coverage: TargetCoverageReport | None = None,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is not None:
        workbook.remove(active_sheet)

    structural_events = [event for event in result.events if is_structural_variant(event)]
    _write_summary(workbook, result)
    _write_key_findings(workbook, structural_events)
    _write_coverage(workbook, target_coverage, selection_coverage)
    _write_qc(workbook, result)
    _write_cnv(workbook, result)
    _write_sv(workbook, structural_events)
    _write_fusions(workbook, structural_events)
    _write_supporting_sheets(workbook, result)

    workbook.save(output_path)
    return output_path
