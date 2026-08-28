from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from ontseq_platform.demo import build_demo_result
from ontseq_platform.io import write_json
from ontseq_platform.models import (
    EventType,
    FusionAnnotation,
    FusionPartnerAnnotation,
    PipelineResult,
)
from ontseq_platform.report import render_html
from ontseq_platform.workbook import render_workbook


class OutputTests(unittest.TestCase):
    def test_demo_round_trip_and_reports(self) -> None:
        result = build_demo_result()
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            json_path = write_json(result, root / "result.json")
            html_path = render_html(result, root / "report.html")
            xlsx_path = render_workbook(result, root / "report.xlsx")

            restored = PipelineResult.model_validate(json.loads(json_path.read_text()))
            self.assertEqual(restored.manifest.sample_id, "SYNTHETIC_AML_001")
            self.assertIn("RESEARCH USE ONLY", html_path.read_text(encoding="utf-8"))

            workbook = load_workbook(xlsx_path, read_only=True)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "00_Summary",
                    "01_QC",
                    "02_CNV_Chromosomes",
                    "03_CNV_Segments",
                    "04_SV",
                    "05_Fusions",
                    "06_ISCN",
                    "07_Warnings",
                    "08_Methods_Versions",
                    "09_Run_Log",
                    "10_Module_Status",
                ],
            )
            workbook.close()

    def test_annotated_translocation_is_included_in_fusion_sheet(self) -> None:
        result = build_demo_result()
        candidate = next(event for event in result.events if event.event_type == EventType.FUSION)
        annotated_bnd = candidate.model_copy(
            update={
                "event_type": EventType.TRANSLOCATION,
                "fusion_evidence": FusionAnnotation(
                    gene_a=FusionPartnerAnnotation(
                        gene="RUNX1T1", preferred_transcript="ENST_RUNX1T1"
                    ),
                    gene_b=FusionPartnerAnnotation(gene="RUNX1", preferred_transcript="ENST_RUNX1"),
                    orientation="+-",
                    frame_status="unknown",
                ),
            }
        )
        result = result.model_copy(
            update={
                "events": [
                    annotated_bnd if event.event_id == candidate.event_id else event
                    for event in result.events
                ]
            }
        )

        with tempfile.TemporaryDirectory() as temporary:
            path = render_workbook(result, Path(temporary) / "report.xlsx")
            workbook = load_workbook(path, read_only=True)
            rows = list(workbook["05_Fusions"].iter_rows(values_only=True))
            workbook.close()

        self.assertEqual(rows[1][0], "FUS-001")
        self.assertEqual(rows[1][6], "+-")


if __name__ == "__main__":
    unittest.main()
