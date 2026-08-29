from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from ontseq_platform.models import (
    AlignedBamIntakeReport,
    AnalysisModule,
    AnalysisSpec,
    AssayMode,
    AssaySpec,
    CheckStatus,
    CraminoQCReport,
    CuteSvCallReport,
    CuteSvPolicy,
    EventType,
    Evidence,
    FileFingerprint,
    FusionAnnotation,
    FusionPartnerAnnotation,
    GenomeBuild,
    GenomicEvent,
    InputKind,
    InputSpec,
    Locus,
    ModuleRunStatus,
    QCMetrics,
    SampleManifest,
    SnifflesCallReport,
    SnifflesPolicy,
    SvConsensusPolicy,
    SvConsensusReport,
    ToolRecord,
    ValidationCheck,
    Verdict,
)
from ontseq_platform.mvp import assemble_aligned_bam_mvp
from ontseq_platform.workbook import render_workbook


def _manifest() -> SampleManifest:
    return SampleManifest(
        sample_id="SYNTHETIC_001",
        run_id="SYNTHETIC_RUN_001",
        input=InputSpec(
            kind=InputKind.ALIGNED_BAM,
            path="/secure/SYNTHETIC_001.bam",
            index_path="/secure/SYNTHETIC_001.bam.bai",
        ),
        assay=AssaySpec(
            mode=AssayMode.LOW_COVERAGE_WGS,
            genome_build=GenomeBuild.GRCH38,
            reference_id="SYNTHETIC_REF",
        ),
        analysis=AnalysisSpec(
            profile="lcwgs",
            modules=[
                AnalysisModule.QC,
                AnalysisModule.CNV,
                AnalysisModule.SV,
                AnalysisModule.ISCN,
                AnalysisModule.REPORT,
            ],
        ),
    )


class AlignedBamMVPTests(unittest.TestCase):
    def test_unrun_callers_are_explicit_not_biological_negatives(self) -> None:
        manifest = _manifest()
        samtools = ToolRecord(name="samtools", version="1.24")
        cramino = ToolRecord(name="cramino", version="1.3.0")
        intake = AlignedBamIntakeReport(
            sample_id=manifest.sample_id,
            reference_id=manifest.assay.reference_id,
            genome_build=manifest.assay.genome_build,
            checks=[
                ValidationCheck(name="synthetic", status=CheckStatus.PASS, message="synthetic")
            ],
            verdict=Verdict.PASS,
            tool=samtools,
        )
        qc = CraminoQCReport(
            sample_id=manifest.sample_id,
            qc=QCMetrics(
                verdict=Verdict.WARN,
                metrics={"mean_coverage_x": 3.0},
                warnings=["No validated numeric gates"],
            ),
            tool=cramino,
        )
        result = assemble_aligned_bam_mvp(
            manifest,
            intake,
            qc,
            pipeline_version="0.2.0-dev",
            git_commit="SYNTHETIC",
        )

        self.assertEqual(result.events, [])
        self.assertEqual(result.iscn.notation, "NOT GENERATED")
        status = {item.module: item.status for item in result.modules}
        self.assertEqual(status[AnalysisModule.QC], ModuleRunStatus.COMPLETED)
        self.assertEqual(status[AnalysisModule.CNV], ModuleRunStatus.NOT_RUN)
        self.assertEqual(status[AnalysisModule.SV], ModuleRunStatus.NOT_RUN)
        self.assertTrue(any("not a biological negative" in item for item in result.iscn.warnings))

    def test_sniffles_candidates_are_visible_but_remain_non_reportable(self) -> None:
        manifest = _manifest()
        samtools = ToolRecord(name="samtools", version="1.24")
        cramino = ToolRecord(name="cramino", version="1.3.0")
        sniffles_tool = ToolRecord(name="Sniffles2", version="2.8.0")
        intake = AlignedBamIntakeReport(
            sample_id=manifest.sample_id,
            reference_id=manifest.assay.reference_id,
            genome_build=manifest.assay.genome_build,
            checks=[],
            verdict=Verdict.PASS,
            tool=samtools,
        )
        qc = CraminoQCReport(
            sample_id=manifest.sample_id,
            qc=QCMetrics(verdict=Verdict.WARN, metrics={}),
            tool=cramino,
        )
        event = GenomicEvent(
            event_id="SNIFFLES2-000001",
            event_type=EventType.DELETION,
            primary=Locus(chromosome="chr1", start=999, end=1200),
            evidence=[Evidence(caller="Sniffles2", caller_version="2.8.0", support_reads=8)],
            reportable=False,
        )
        sniffles = SnifflesCallReport(
            sample_id=manifest.sample_id,
            genome_build=manifest.assay.genome_build,
            status=ModuleRunStatus.COMPLETED,
            policy=SnifflesPolicy(
                profile_id="synthetic",
                status="technical_defaults_only",
                note="Not clinically validated.",
            ),
            events=[event],
            raw_record_count=1,
            accepted_record_count=1,
            rejected_record_count=0,
            tool=sniffles_tool,
            vcf_fingerprint=FileFingerprint(size_bytes=100, sha256="a" * 64),
        )

        result = assemble_aligned_bam_mvp(
            manifest,
            intake,
            qc,
            pipeline_version="0.3.0-dev",
            git_commit="SYNTHETIC",
            sniffles_report=sniffles,
        )

        self.assertEqual([item.event_id for item in result.events], [event.event_id])
        self.assertEqual(result.events[0].confidence, "low")
        self.assertFalse(result.events[0].reportable)
        status = {item.module: item.status for item in result.modules}
        self.assertEqual(status[AnalysisModule.SV], ModuleRunStatus.COMPLETED)
        self.assertEqual(status[AnalysisModule.ISCN], ModuleRunStatus.NOT_RUN)
        self.assertEqual(result.iscn.notation, "NOT GENERATED")
        self.assertEqual(result.provenance.reference_checksums["sniffles_vcf"], "a" * 64)
        with tempfile.TemporaryDirectory() as directory:
            path = render_workbook(result, Path(directory) / "result.xlsx")
            workbook = load_workbook(path, read_only=True)
            cnv_ids = {row[0] for row in workbook["03_CNV_Segments"].iter_rows(values_only=True)}
            sv_rows = list(workbook["04_SV"].iter_rows(values_only=True))
            workbook.close()
        self.assertNotIn(event.event_id, cnv_ids)
        self.assertEqual(sv_rows[1][0], event.event_id)
        reportable_index = sv_rows[0].index("reportable")
        self.assertFalse(sv_rows[1][reportable_index])

    def test_fusion_module_reports_completed_breakpoint_assessment(self) -> None:
        manifest = _manifest()
        intake = AlignedBamIntakeReport(
            sample_id=manifest.sample_id,
            reference_id=manifest.assay.reference_id,
            genome_build=manifest.assay.genome_build,
            checks=[],
            verdict=Verdict.PASS,
            tool=ToolRecord(name="samtools", version="1.24"),
        )
        qc = CraminoQCReport(
            sample_id=manifest.sample_id,
            qc=QCMetrics(verdict=Verdict.WARN, metrics={}),
            tool=ToolRecord(name="cramino", version="1.3.0"),
        )
        event = GenomicEvent(
            event_id="SVCLUSTER-000001",
            event_type=EventType.TRANSLOCATION,
            primary=Locus(chromosome="chr10", start=21_634_899, end=21_634_900),
            secondary=Locus(chromosome="chr11", start=85_975_045, end=85_975_046),
            evidence=[Evidence(caller="cuteSV", caller_version="2.1.3", support_reads=39)],
            confidence="high",
            known_rearrangement="PICALM::MLLT10",
            fusion_evidence=FusionAnnotation(
                gene_a=FusionPartnerAnnotation(gene="PICALM"),
                gene_b=FusionPartnerAnnotation(gene="MLLT10"),
            ),
        )
        cutesv = CuteSvCallReport(
            sample_id=manifest.sample_id,
            genome_build=manifest.assay.genome_build,
            status=ModuleRunStatus.COMPLETED,
            policy=CuteSvPolicy(
                profile_id="synthetic",
                status="technical_defaults_only",
                note="Synthetic technical defaults.",
            ),
            events=[event],
            raw_record_count=1,
            accepted_record_count=1,
            rejected_record_count=0,
            tool=ToolRecord(name="cuteSV", version="2.1.3"),
            vcf_fingerprint=FileFingerprint(size_bytes=100, sha256="b" * 64),
        )
        consensus = SvConsensusReport(
            sample_id=manifest.sample_id,
            genome_build=manifest.assay.genome_build,
            status=ModuleRunStatus.COMPLETED,
            policy=SvConsensusPolicy(
                profile_id="synthetic",
                status="technical_defaults_only",
                note="Synthetic technical defaults.",
            ),
            events=[event],
            input_event_count=1,
            consolidated_event_count=1,
            caller_names=["cuteSV"],
        )

        result = assemble_aligned_bam_mvp(
            manifest,
            intake,
            qc,
            pipeline_version="0.5.3",
            git_commit="SYNTHETIC",
            cutesv_report=cutesv,
            sv_consensus_report=consensus,
        )

        fusion_module = next(
            item for item in result.modules if item.module == AnalysisModule.FUSION
        )
        self.assertEqual(fusion_module.status, ModuleRunStatus.COMPLETED)
        self.assertIn("1 event(s) carried fusion evidence", fusion_module.reason)
        self.assertIn("1 matched a hematology review pattern", fusion_module.reason)
        self.assertEqual(result.provenance.reference_checksums["cutesv_vcf"], "b" * 64)

    def test_failed_intake_cannot_be_assembled(self) -> None:
        manifest = _manifest()
        intake = AlignedBamIntakeReport(
            sample_id=manifest.sample_id,
            reference_id=manifest.assay.reference_id,
            genome_build=manifest.assay.genome_build,
            checks=[
                ValidationCheck(name="synthetic", status=CheckStatus.FAIL, message="synthetic")
            ],
            verdict=Verdict.FAIL,
        )
        qc = CraminoQCReport(
            sample_id=manifest.sample_id,
            qc=QCMetrics(verdict=Verdict.WARN, metrics={}),
            tool=ToolRecord(name="cramino", version="synthetic"),
        )
        with self.assertRaises(ValueError):
            assemble_aligned_bam_mvp(
                manifest,
                intake,
                qc,
                pipeline_version="0.2.0-dev",
                git_commit="SYNTHETIC",
            )


if __name__ == "__main__":
    unittest.main()
