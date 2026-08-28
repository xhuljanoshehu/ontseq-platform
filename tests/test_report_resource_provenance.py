from __future__ import annotations

import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from ontseq_platform.demo import build_demo_result
from ontseq_platform.models import (
    GenomeBuild,
    ReferenceDictionaryContract,
    ResolvedResourceContext,
)
from ontseq_platform.report import render_html
from ontseq_platform.workbook import render_workbook


class ReportResourceProvenanceTests(unittest.TestCase):
    def test_resource_releases_and_checksums_render_in_methods_provenance(self) -> None:
        with TemporaryDirectory() as raw:
            root = Path(raw).resolve()
            resource = root / "reference.fa"
            resource.write_text(">chr1\nA\n", encoding="utf-8")
            context = ResolvedResourceContext(
                profile_id="AML_LCWGS_GRCh38",
                profile_version="v1",
                genome_build=GenomeBuild.GRCH38,
                reference_dictionary_contract=(ReferenceDictionaryContract.GRCH38_CANONICAL_25),
                reference_bundle_id="GRCh38_GENCODE50_MANE1.5_v1",
                reference_bundle_version="v1",
                knowledge_bundle_id="HEMATOLOGY_v1",
                knowledge_bundle_version="v1",
                resource_root=str(root),
                resource_paths={"reference.genome_fasta": str(resource)},
                resource_checksums={"reference.genome_fasta": "a" * 64},
                resource_releases={
                    "reference.genome_fasta": "GRCh38.p14",
                    "reference.gencode_gtf": "GENCODE 50",
                    "reference.mane_gff3": "MANE 1.5",
                    "reference.cytobands": "UCSC hg38",
                },
            )
            result = build_demo_result().model_copy(update={"reference_context": context})
            html_path = render_html(result, root / "report.html")
            html = html_path.read_text(encoding="utf-8")
            for expected in (
                "GRCh38.p14",
                "GRCh38_GENCODE50_MANE1.5_v1",
                "grch38_canonical_25",
                "GENCODE 50",
                "MANE 1.5",
                "UCSC hg38",
                "HEMATOLOGY_v1",
                "Resource SHA256 provenance",
            ):
                self.assertIn(expected, html)

            workbook_path = render_workbook(result, root / "report.xlsx")
            workbook = load_workbook(workbook_path, read_only=True)
            try:
                values = {
                    row[0].value: row[1].value
                    for row in workbook["00_Summary"].iter_rows(min_row=2)
                    if row[0].value is not None
                }
                self.assertEqual(values["Genome assembly"], "GRCh38.p14")
                self.assertIn("GRCh38_GENCODE50", values["ReferenceBundle"])
                self.assertEqual(
                    values["BAM dictionary contract"],
                    "grch38_canonical_25",
                )
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
