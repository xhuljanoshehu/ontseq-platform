"""Derive the versioned Adaptive Sampling panel and its coverage expectations.

The two source workbooks are laboratory records. They are **not** part of this repository
and must not be: they carry sample identifiers. This script reads them from a path the
operator supplies, and writes only de-identified derivatives:

* the panel intervals, which are genomic coordinates and carry no sample information;
* per-target coverage expectations, aggregated across runs with the run labels replaced by
  ordinal positions.

Running it is therefore reproducible without the raw workbooks ever entering Git. The
committed outputs record the SHA256 of each source so a later reviewer can prove which
revision they came from.

Usage::

    python scripts/build_adaptive_sampling_panel.py \
        --coverages /path/to/AS_coverages_fusions.xlsx \
        --experiments /path/to/ONT_experiments_results.xlsx \
        --panel-version AS_FUSION_PANEL_V1_UNCONFIRMED
"""

from __future__ import annotations

import argparse
import hashlib
import statistics
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook

#: GRCh38 primary assembly lengths, used to prove every interval is in range for the build
#: the panel claims. Taken from the assembly report, not from any run output.
GRCH38_LENGTHS: dict[str, int] = {
    "chr1": 248956422,
    "chr2": 242193529,
    "chr3": 198295559,
    "chr4": 190214555,
    "chr5": 181538259,
    "chr6": 170805979,
    "chr7": 159345973,
    "chr8": 145138636,
    "chr9": 138394717,
    "chr10": 133797422,
    "chr11": 135086622,
    "chr12": 133275309,
    "chr13": 114364328,
    "chr14": 107043718,
    "chr15": 101991189,
    "chr16": 90338345,
    "chr17": 83257441,
    "chr18": 80373285,
    "chr19": 58617616,
    "chr20": 64444167,
    "chr21": 46709983,
    "chr22": 50818468,
    "chrX": 156040895,
    "chrY": 57227415,
}

#: Rows whose gene label and coordinates disagree with the public annotation. They are kept
#: in the output so nothing is silently dropped, but they are named here and in the lock so
#: that a reviewer has to resolve them before the panel is promoted.
OPEN_QUESTIONS: dict[str, str] = {
    "IGH": (
        "Labelled IGH but placed at chr5:143,396,959-143,417,420. IGH is located on "
        "chr14q32 in GRCh38. Either the label or the interval is wrong; both source "
        "workbooks carry the same value, so the error is upstream of this repository."
    ),
}


@dataclass(frozen=True)
class Target:
    gene: str
    chromosome: str
    start: int
    end: int

    @property
    def region(self) -> str:
        return f"{self.chromosome}:{self.start}-{self.end}"

    @property
    def length(self) -> int:
        return self.end - self.start


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def read_targets(workbook: Path) -> list[Target]:
    """Read the panel from the coverage workbook and check it against its own region column."""
    sheet = load_workbook(workbook, data_only=True)["coverages (2)"]
    targets: list[Target] = []
    for index in range(3, sheet.max_row + 1):
        gene = sheet.cell(row=index, column=2).value
        chromosome = sheet.cell(row=index, column=3).value
        start = sheet.cell(row=index, column=4).value
        end = sheet.cell(row=index, column=5).value
        region = sheet.cell(row=index, column=7).value
        if not gene or not chromosome or start is None or end is None:
            continue
        target = Target(str(gene).strip(), str(chromosome).strip(), int(start), int(end))
        if region and str(region).strip() != target.region:
            raise SystemExit(f"{target.gene}: region string {region!r} contradicts the columns")
        targets.append(target)
    return targets


def read_expectations(workbook: Path, targets: list[Target]) -> dict[str, list[float]]:
    """Per-target observed coverage across runs, keyed by region, run labels discarded."""
    sheet = load_workbook(workbook, data_only=True)["Coverage_per_gene"]
    known = {target.region for target in targets}
    observed: dict[str, list[float]] = {}
    for index in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=index, column=1).value
        region = str(cell).strip() if cell else None
        if region is None or region not in known:
            continue
        values = [
            float(sheet.cell(row=index, column=column).value)
            for column in range(4, sheet.max_column + 1)
            if isinstance(sheet.cell(row=index, column=column).value, int | float)
        ]
        observed[region] = values
    return observed


def validate(targets: list[Target]) -> list[str]:
    """Every check that can be made without leaving this repository. Failing one is fatal."""
    problems: list[str] = []
    seen_regions: set[str] = set()
    seen_genes: set[str] = set()
    for target in targets:
        if target.end <= target.start:
            problems.append(f"{target.gene}: end is not greater than start")
        limit = GRCH38_LENGTHS.get(target.chromosome)
        if limit is None:
            problems.append(f"{target.gene}: {target.chromosome} is not a GRCh38 primary contig")
        elif target.end > limit:
            problems.append(f"{target.gene}: {target.region} exceeds the GRCh38 contig length")
        if target.region in seen_regions:
            problems.append(f"{target.region}: duplicate interval")
        if target.gene in seen_genes:
            problems.append(f"{target.gene}: duplicate gene label")
        seen_regions.add(target.region)
        seen_genes.add(target.gene)
    ordered = sorted(targets, key=lambda item: (item.chromosome, item.start))
    for left, right in zip(ordered, ordered[1:], strict=False):
        if left.chromosome == right.chromosome and right.start < left.end:
            problems.append(f"{left.gene} and {right.gene} overlap")
    return problems


def write_bed(path: Path, targets: list[Target], *, panel_version: str) -> None:
    ordered = sorted(
        targets,
        key=lambda item: (
            int(item.chromosome[3:]) if item.chromosome[3:].isdigit() else 99,
            item.chromosome,
            item.start,
        ),
    )
    lines = [
        "# ONTSeq Adaptive Sampling panel - RESEARCH USE ONLY, NOT A VALIDATED ASSAY DESIGN",
        f"# genome_build=GRCh38 panel_version={panel_version} targets={len(ordered)}",
        "# Coordinates are reproduced verbatim from the laboratory source. The design carries",
        "# approximately 10 kb flanks around each gene, so this is a BUFFERED panel and must",
        "# not be used where an unbuffered analysis ROI is required.",
    ]
    for target in ordered:
        name = target.gene
        if target.gene in OPEN_QUESTIONS:
            name = f"{target.gene}_REVIEW_REQUIRED"
        lines.append(f"{target.chromosome}\t{target.start}\t{target.end}\t{name}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_lock(
    path: Path,
    *,
    bed_path: Path,
    panel_version: str,
    targets: list[Target],
    sources: dict[str, str],
) -> None:
    chromosomes = sorted(
        {target.chromosome for target in targets},
        key=lambda name: (int(name[3:]) if name[3:].isdigit() else 99, name),
    )
    document = {
        "schema_version": "0.1.0",
        "panel_version": panel_version,
        "status": "derived_unconfirmed",
        "genome_build": "GRCh38",
        "role": "selection_panel_buffered",
        "generated_at": datetime.now(UTC).strftime("%Y-%m-%d"),
        "bed": {
            "path": str(bed_path.as_posix()),
            "sha256": sha256_file(bed_path),
            "target_count": len(targets),
            "interval_bases": sum(target.length for target in targets),
            "chromosomes": chromosomes,
        },
        "derived_from": sources,
        "evidence": [
            "Both laboratory workbooks describe an identical set of intervals.",
            "The chromosome set matches the contigs emitted by the legacy Sniffles run, "
            "which was restricted with --regions fusion_panel_with_buffer.bed.",
            "Interval ends sit exactly 10,000 bp beyond the Ensembl GRCh38 gene end for the "
            "targets checked, which is what makes this a buffered design.",
        ],
        "open_questions": [
            {"target": gene, "detail": detail} for gene, detail in sorted(OPEN_QUESTIONS.items())
        ],
        "not_established": [
            "That this file is byte-identical to the panel the sequencer selected on.",
            "That the coordinate convention is 0-based half-open rather than 1-based inclusive.",
            "Any coverage, reportability or no-call threshold for this design.",
        ],
    }
    path.write_text(yaml.safe_dump(document, sort_keys=False, allow_unicode=True), encoding="utf-8")


def write_expectations(
    path: Path, targets: list[Target], observed: dict[str, list[float]], *, panel_version: str
) -> None:
    header = [
        "# Observed per-target mean coverage across historical Adaptive Sampling runs.",
        "# DESCRIPTIVE ONLY. These are not adequacy gates, not reportability thresholds and",
        "# not a no-call definition. Run labels are deliberately reduced to a count so that",
        "# no sample can be identified from this file.",
        f"# genome_build=GRCh38 panel_version={panel_version}",
        "gene\tchromosome\tstart\tend\tlength\truns\tmin_mean_depth\tmedian_mean_depth\tmax_mean_depth",
    ]
    lines = list(header)
    for target in sorted(targets, key=lambda item: item.gene):
        values = [value for value in observed.get(target.region, []) if value > 0]
        if not values:
            continue
        lines.append(
            "\t".join(
                [
                    target.gene,
                    target.chromosome,
                    str(target.start),
                    str(target.end),
                    str(target.length),
                    str(len(values)),
                    f"{min(values):.2f}",
                    f"{statistics.median(values):.2f}",
                    f"{max(values):.2f}",
                ]
            )
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--coverages", type=Path, required=True)
    parser.add_argument("--experiments", type=Path, required=True)
    parser.add_argument("--panel-version", default="AS_FUSION_PANEL_V1_UNCONFIRMED")
    parser.add_argument("--output-dir", type=Path, default=Path("configs/panels"))
    parser.add_argument("--expectations", type=Path, default=Path("configs/qc"))
    args = parser.parse_args()

    targets = read_targets(args.coverages)
    problems = validate(targets)
    if problems:
        raise SystemExit("panel validation failed:\n  " + "\n  ".join(problems))

    args.output_dir.mkdir(parents=True, exist_ok=True)
    args.expectations.mkdir(parents=True, exist_ok=True)
    stem = "aml_fusion_adaptive_sampling.grch38.buffered"
    bed_path = args.output_dir / f"{stem}.bed"
    write_bed(bed_path, targets, panel_version=args.panel_version)
    write_lock(
        args.output_dir / f"{stem}.lock.yaml",
        bed_path=bed_path,
        panel_version=args.panel_version,
        targets=targets,
        sources={
            "coverage_workbook_sha256": sha256_file(args.coverages),
            "experiment_workbook_sha256": sha256_file(args.experiments),
        },
    )
    write_expectations(
        args.expectations / "target_coverage_expectations.grch38.tsv",
        targets,
        read_expectations(args.experiments, targets),
        panel_version=args.panel_version,
    )
    print(f"{len(targets)} targets written to {bed_path}")


if __name__ == "__main__":
    main()
